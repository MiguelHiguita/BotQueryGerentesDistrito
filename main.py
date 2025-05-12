import os
import datetime
import pandas as pd
import win32com.client
from openpyxl import load_workbook
from collections import defaultdict
from openpyxl.worksheet.table import Table, TableStyleInfo

RUTA_CARPETA = r"T:\AUDITORIA TIENDAS\CONCILIACION TIENDAS\AUTOMATIZACIONES\bot_query_gerentes_distrito\archivos"
RUTA_TEMPORAL = r"T:\AUDITORIA TIENDAS\CONCILIACION TIENDAS\AUTOMATIZACIONES\bot_query_gerentes_distrito\archivos_temporales"
RUTA_BD_GERENTES = r"T:\AUDITORIA TIENDAS\CONCILIACION TIENDAS\AUTOMATIZACIONES\bot_query_gerentes_distrito\base_datos\BASE DE DATOS....xlsx"
RUTA_HISTORICO = r"T:\AUDITORIA TIENDAS\CONCILIACION TIENDAS\AUTOMATIZACIONES\bot_query_gerentes_distrito\historico"

REMITENTES = {
    "marianagl": "Mariana Giraldo Lopez",
    "andreina": "Andreina Peralta Oliveros",
    "conciliacion": "Estefanía Ramirez",
    "luisasg": "Luisa Fernanda Sánchez Garcia",
    "amsanchez": "Ana Maria Sanchez Tamayo",
    "Aprendizps1": "Miguel Angel Gonzalez Higuita"
}

def limpiar_archivos_temporales():
    archivos = [f for f in os.listdir(RUTA_TEMPORAL) if os.path.isfile(os.path.join(RUTA_TEMPORAL, f))]
    for archivo in archivos:
        os.remove(rf"{RUTA_TEMPORAL}\{archivo}")

def enviar_correo(outlook, destinatario, adjuntos):
    fecha_actual = datetime.datetime.now().strftime("%d/%m/%Y")
    hora_actual = datetime.datetime.now().strftime("%I:%M %p").lower()
    mail = None
    try:
        usuario = os.getlogin()
        remitente = REMITENTES.get(usuario, "Equipo")
        mail = outlook.CreateItem(0)
        mail.To = destinatario
        # mail.CC = "analiticacontraloria@gco.com.co"
        mail.Subject = f"COMPRAS Y TRANSFERENCIAS PENDIENTES POR ACTUALIZAR TIENDAS FRANQUICIAS Y PROPIAS {fecha_actual}"
        mail.HTMLbody = f"""
        <p>Buen día;</p>
        <p>¡Espero y se encuentren muy bien!</p>
        <p>Adjunto informe con las compras y transferencias pendientes por actualizar en Y2 hasta el cierre del día de hoy
        {fecha_actual} a las {hora_actual}, para las tiendas a nivel nacional. Éste se puede modificar y colocar la marca o
        tipo tienda. La gestión se debe realizar de la siguiente manera:</p>
        <ul>
            <li>Compras pendientes por actualizar: Los documentos que tienen un rango de 1 a 5 días vencidos se
            encuentran en transito y están próximos a la entrega, los que tengan más de 6 días
            se debe verificar con la tienda para que gestione las guías de transporte (tienda-CEDI).</li>
            <li>Transferencias de tienda a tienda: Los documentos pendientes por recibir con rango mayor a 5 días 
            vencidos, deben gestionarse con la tienda origen solicitando el número de guía. Si pasados 2 días 
            calendario no se recibe respuesta del origen con el número de guía, actualizar el documento y cargarlo a 
            la tienda origen notificando sobre el mismo correo con copia al gerente de distrito, jefe de zona, 
            coordinador o supervisor.</li>
        </ul>
        <p>Quedo atenta;</p>
        <p>{remitente}</p>
        """
        if isinstance(adjuntos, str):
            adjuntos = [adjuntos]
        for adjunto in adjuntos:
            if os.path.exists(adjunto):
                mail.Attachments.Add(adjunto)
            else:
                print(f"El archivo {adjunto} no existe")
        mail.Send()
        print(f"Correo enviado a {destinatario}")
    except Exception as e:
        print(f"Error al enviar el correo a {destinatario}: {e}")
    finally:
        if mail:
            mail = None

def separar_archivos_por_gerente():
    data_gerentes = pd.read_excel(RUTA_BD_GERENTES)
    data_gerentes = data_gerentes.dropna(subset=['POS'])
    data_gerentes['POS'] = data_gerentes['POS'].fillna(0).astype(float).astype(int)
    pos_to_correo = dict(zip(data_gerentes['POS'], data_gerentes['CORREO CORPORATIVO']))
    archivos_por_gerente = defaultdict(list)
    outlook = win32com.client.Dispatch("Outlook.Application")
    outlook.GetNamespace("MAPI").GetDefaultFolder(6).Display()
    destinatario = "aprendizprogramacion1@gco.com.co"
    for archivo in os.listdir(RUTA_TEMPORAL):
        if archivo.endswith(".xlsx"):
            pos = archivo.split("_")[0]
            try:
                pos_int = int(pos)
                if pos_int in pos_to_correo:
                    correo_gerente = pos_to_correo[pos_int]
                    ruta_completa = os.path.join(RUTA_TEMPORAL, archivo)
                    archivos_por_gerente[correo_gerente].append(ruta_completa)
            except ValueError:
                print(f"Error al procesar el archivo {archivo}: POS no válido")
    for correo_gerente, archivos in archivos_por_gerente.items():
        try:
            enviar_correo(outlook, destinatario, archivos)
        except Exception as e:
            print(f"Error al enviar correo a {correo_gerente}: {e}")
    limpiar_archivos_temporales()

def mover_archivo(ruta_origen, ruta_destino):
    if not os.path.exists(ruta_destino):
        os.makedirs(ruta_destino)
    nombre_archivo = os.path.basename(ruta_origen)
    ruta_destino_completa = os.path.join(ruta_destino, nombre_archivo)
    contador = 1
    nombre_archivo_sin_ext, extension = os.path.splitext(nombre_archivo)
    while os.path.exists(ruta_destino_completa):
        ruta_destino_completa = os.path.join(ruta_destino, f"{nombre_archivo_sin_ext}_{contador}{extension}")
        contador += 1
    os.rename(ruta_origen, ruta_destino_completa)

def obtener_rutas(ruta_carpeta):
    archivos = [os.path.join(ruta_carpeta, archivo) for archivo in os.listdir(ruta_carpeta) if os.path.isfile(os.path.join(ruta_carpeta, archivo))]
    ruta_primer_archivo = archivos[0] if len(archivos) > 0 else None
    ruta_segundo_archivo = archivos[1] if len(archivos) > 1 else None
    return ruta_primer_archivo, ruta_segundo_archivo

def generar_tabla_xlsx(ruta_archivo):
    wb = load_workbook(ruta_archivo)
    ws = wb.active
    max_col, max_row = ws.max_column, ws.max_row
    table_range = f"{ws.cell(row=1, column=1).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}"
    tabla = Table(displayName="Tabla1", ref=table_range)
    estilo = TableStyleInfo(name="TableStyleLight13", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)
    wb.save(ruta_archivo)

def generar_excels_temporales(ruta_archivo):
    data = pd.read_excel(ruta_archivo, parse_dates=False)
    data = data[~data["Nombre Tienda"].str.contains("Chat", na=False)]
    mango_mask = (data["Nombre Tienda"].str.contains("Mng|Mango", case=False, na=False)) & (~data["Nombre Tienda"].str.contains("Outlet", case=False, na=False))
    data = data[~mango_mask]
    nombre_base_archivo = "Transferencias.xlsx" if "Pos Destino" in data.columns else "Compras.xlsx"
    if not os.path.exists(RUTA_TEMPORAL):
        os.makedirs(RUTA_TEMPORAL)
    for pos in data["Pos Destino"].unique() if nombre_base_archivo == "Transferencias.xlsx" else data["Pos"].unique():
        if pos is not None:
            data_filtrada = data[data["Pos Destino"] == pos] if nombre_base_archivo == "Transferencias.xlsx" else data[data["Pos"] == pos]
            temp_inicial = os.path.join(RUTA_TEMPORAL, f"{pos}_{nombre_base_archivo}")
            data_filtrada.to_excel(temp_inicial, index=False)
            generar_tabla_xlsx(temp_inicial)
            mover_archivo(temp_inicial, RUTA_TEMPORAL)

def main():
    ruta_primer_archivo, ruta_segundo_archivo = obtener_rutas(RUTA_CARPETA)
    if ruta_primer_archivo and ruta_segundo_archivo:
        generar_excels_temporales(ruta_primer_archivo)
        generar_excels_temporales(ruta_segundo_archivo)
        separar_archivos_por_gerente()
        mover_archivo(ruta_primer_archivo, RUTA_HISTORICO)
        mover_archivo(ruta_segundo_archivo, RUTA_HISTORICO)
    elif ruta_primer_archivo:
        generar_excels_temporales(ruta_primer_archivo)
        separar_archivos_por_gerente()
        mover_archivo(ruta_primer_archivo, RUTA_HISTORICO)
    else:
        print("No se encontraron archivos en la carpeta")

if __name__ == "__main__":
    main()