# ANTES DE EJECUTAR ESTE CÓDIGO, SE DEBE VERIFICAR SI LA FUNCIÓN LLAMADA EN "generar_y_enviar_excels_temporales" 
# PARA ENVIAR LOS CORREOS ES "enviar_correo" O "prueba_enviar_correo"
# EN CASO DE SER "enviar_correo" LOS CORREOS SE ENVIARÁN A LOS GERENTES Y SE AGREGARÁ EN COPIA A TODO EL EQUIPO DE ANALITICA CONTRALORÍA
# POR OTRO LADO "prueba_enviar_correo" SIRVE PARA HACER PRUEBAS, SE ENVÍAN TODOS LOS CORREOS A MIGUEL Y NO SE AGREGA A NADIE EN COPIA

import os
import re
import datetime
import pandas as pd
import win32com.client
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

RUTA_CARPETA = r"T:\AUDITORIA TIENDAS\CONCILIACION TIENDAS\AUTOMATIZACIONES\bot_query_gerentes_distrito\archivo"
RUTA_CONCATENADO = r"T:\AUDITORIA TIENDAS\CONCILIACION TIENDAS\AUTOMATIZACIONES\bot_query_gerentes_distrito\concatenado"
RUTA_TEMPORAL = r"T:\AUDITORIA TIENDAS\CONCILIACION TIENDAS\AUTOMATIZACIONES\bot_query_gerentes_distrito\temporal"
RUTA_BD_GERENTES = r"T:\AUDITORIA TIENDAS\CONCILIACION TIENDAS\AUTOMATIZACIONES\Correos Gerentes\BASE DE DATOS....xlsx"
RUTA_HISTORICO = r"T:\AUDITORIA TIENDAS\CONCILIACION TIENDAS\AUTOMATIZACIONES\bot_query_gerentes_distrito\historico"

REMITENTES = {
    "marianagl": "Mariana Giraldo Lopez",
    "andreina": "Andreina Peralta Oliveros",
    "conciliacion": "Estefanía Ramirez",
    "luisasg": "Luisa Fernanda Sánchez Garcia",
    "amsanchez": "Ana Maria Sanchez Tamayo",
    "Aprendizps1": "Miguel Angel Gonzalez Higuita"
}

def limpiar_carpeta(carpeta):
    archivos = [f for f in os.listdir(carpeta) if os.path.isfile(os.path.join(carpeta, f))]
    for archivo in archivos:
        os.remove(rf"{carpeta}\{archivo}")

def enviar_correo(outlook, destinatario, adjunto):
    fecha_actual = datetime.datetime.now().strftime("%d/%m/%Y")
    hora_actual = datetime.datetime.now().strftime("%I:%M %p").lower()
    usuario = os.getlogin()
    remitente = REMITENTES.get(usuario, "Equipo")
    try:
        if destinatario == "mariana.giraldo@gco.com.co":
            try:
                mail = outlook.CreateItem(0)
                mail.Subject = f"⚠️⚠️⚠️COMPRAS Y TRANSFERENCIAS PENDIENTES POR ACTUALIZAR TIENDAS FRANQUICIAS Y PROPIAS {fecha_actual}⚠️⚠️⚠️"
                mail.HTMLbody = f"""
                MARIANAAAAA, validar estas compras y transferencias pendientes ya que hay tiendas que no tienen un correo asignado en la base de datos
                """
                mail.To = destinatario
                mail.CC = "conciliaciontiendas.pr@gco.com.co"
                mail.Attachments.Add(adjunto)
                mail.Send()
                print(f"""
                Correo de {destinatario} enviado con exito.
                """)
            except Exception as e:
                print(f"Error al enviar el correo de {destinatario}: {e}")
            finally:
                if mail:
                    mail = None
        else:
            try:
                mail = outlook.CreateItem(0)
                mail.Subject = f"COMPRAS Y TRANSFERENCIAS PENDIENTES POR ACTUALIZAR TIENDAS FRANQUICIAS Y PROPIAS {fecha_actual}"
                mail.HTMLbody = f"""
                <p>Buen día</p>
                <p>¡Espero y se encuentre muy bien!</p>
                <p>Adjunto informe con las compras y transferencias pendientes por actualizar en Y2 hasta el cierre del día de hoy
                {fecha_actual} a las {hora_actual}, para las tiendas a nivel nacional. Éste se puede modificar y colocar la marca o
                tipo tienda. La gestión se debe realizar de la siguiente manera:</p>
                <ul>
                    <li>Compras pendientes por actualizar: Los documentos que tienen un rango de 1 a 5 días vencidos se
                    encuentran en transito y están próximos a la entrega, los que tengan más de 6 días
                    se debe verificar con la tienda para que gestione las guías de transporte (tienda-CEDI).</li>
                    <li>Transferencias de tienda a tienda: Los documentos pendientes por recibir con rango mayor a 5 días 
                    vencidos, deben gestionarse con la tienda origen solicitando el número de guía. Si pasados 2 días 
                    calendario no se recibe respuesta del origen con el número de guía, actualizar el documento y cargarlo a 
                    la tienda origen notificando sobre el mismo correo con copia al gerente de distrito, jefe de zona, 
                    coordinador o supervisor.</li>
                </ul>
                <p>Quedo atenta;</p>
                <p>{remitente}</p>
                """
                mail.To = destinatario  # SE ENVIA A LOS GERENTES DE LAS TIENDAS
                mail.CC = "analiticacontraloria@gco.com.co" # SE PONE EN COPIA A TODO EL EQUIPO ANALITICA CONTRALORIA
                mail.Attachments.Add(adjunto)
                mail.Send()
                print(f"""
                Correo de {destinatario} enviado exitosamente.
                """)
            except Exception as e:
                print(f"Error al enviar el correo de {destinatario}: {e}")
            finally:
                if mail:
                    mail = None
    except Exception as e:
        print(f"Error al enviar los correos {e}")
    finally:
        limpiar_carpeta(RUTA_CONCATENADO)
        limpiar_carpeta(RUTA_TEMPORAL)

def prueba_enviar_correo(outlook, destinatario, adjunto):
    destinatario_prueba = "aprendizprogramacion1@gco.com.co"
    nombre_gerente = destinatario.split('@')[0].replace('.', ' ')
    fecha_actual = datetime.datetime.now().strftime("%d/%m/%Y")
    hora_actual = datetime.datetime.now().strftime("%I:%M %p").lower()
    usuario = os.getlogin()
    remitente = REMITENTES.get(usuario, "Equipo")
    try:
        if destinatario == "mariana.giraldo@gco.com.co":
            try: 
                mail = outlook.CreateItem(0)
                mail.Subject = f"⚠️⚠️⚠️COMPRAS Y TRANSFERENCIAS PENDIENTES POR ACTUALIZAR TIENDAS FRANQUICIAS Y PROPIAS {fecha_actual}⚠️⚠️⚠️"
                mail.HTMLbody = f"""
                MARIANAAAAA, validar estas compras y transferencias pendientes ya que hay tiendas que no tienen un correo asignado en la base de datos
                """
                mail.To = destinatario_prueba
                mail.CC = "mariana.giraldo@gco.com.co; conciliaciontiendas.pr@gco.com.co"
                mail.Attachments.Add(adjunto)
                mail.Send()
                print(f"""
                Correo de: {destinatario}
                Enviado a: {destinatario_prueba} para pruebas
                """)
            except Exception as e:
                print(f"Error al enviar el correo de {destinatario}: {e}")
            finally:
                if mail:
                    mail = None
        else:
            try:
                mail = outlook.CreateItem(0)
                mail.Subject = f"COMPRAS Y TRANSFERENCIAS PENDIENTES POR ACTUALIZAR TIENDAS FRANQUICIAS Y PROPIAS {fecha_actual}"
                mail.HTMLbody = f"""
                <p>Buen día {nombre_gerente};</p>
                <p>¡Espero y se encuentre muy bien!</p>
                <p>Adjunto informe con las compras y transferencias pendientes por actualizar en Y2 hasta el cierre del día de hoy
                {fecha_actual} a las {hora_actual}, para las tiendas a nivel nacional. Éste se puede modificar y colocar la marca o
                tipo tienda. La gestión se debe realizar de la siguiente manera:</p>
                <ul>
                    <li>Compras pendientes por actualizar: Los documentos que tienen un rango de 1 a 5 días vencidos se
                    encuentran en transito y están próximos a la entrega, los que tengan más de 6 días
                    se debe verificar con la tienda para que gestione las guías de transporte (tienda-CEDI).</li>
                    <li>Transferencias de tienda a tienda: Los documentos pendientes por recibir con rango mayor a 5 días 
                    vencidos, deben gestionarse con la tienda origen solicitando el número de guía. Si pasados 2 días 
                    calendario no se recibe respuesta del origen con el número de guía, actualizar el documento y cargarlo a 
                    la tienda origen notificando sobre el mismo correo con copia al gerente de distrito, jefe de zona, 
                    coordinador o supervisor.</li>
                </ul>
                <p>Quedo atent@;</p>
                <p>{remitente}</p>
                """
                mail.To = destinatario_prueba
                mail.Attachments.Add(adjunto)
                mail.Send()
                print(f"""
                Correo de: {destinatario}
                Enviado a: {destinatario_prueba} para pruebas
                """)
            except Exception as e:
                print(f"Error al enviar el correo de {destinatario}: {e}")
            finally:
                if mail:
                    mail = None
    except Exception as e:
        print(f"Error al enviar los correos {e}")
    finally:
        limpiar_carpeta(RUTA_CONCATENADO)
        limpiar_carpeta(RUTA_TEMPORAL)

def eliminar_columna(worksheet, columna_a_eliminar):
    for idx, col in enumerate(worksheet[1], 1):
        if col.value == columna_a_eliminar:
            worksheet.delete_cols(idx, 1)
            break

def mover_archivo(ruta_origen, ruta_destino):
    try:
        if not os.path.exists(ruta_destino):
            os.makedirs(ruta_destino)
        nombre_archivo = os.path.basename(ruta_origen)
        nombre_archivo_sin_ext, extension = os.path.splitext(nombre_archivo)
        ruta_destino_completa = os.path.join(ruta_destino, nombre_archivo)
        contador = 1
        while os.path.exists(ruta_destino_completa):
            ruta_destino_completa = os.path.join(ruta_destino, f"{nombre_archivo_sin_ext}_{contador}{extension}")
            contador += 1
        os.rename(ruta_origen, ruta_destino_completa)
        print(f"Archivo movido exitosamente: {os.path.basename(ruta_destino_completa)}")
    except Exception as e:
        print(f"Error al mover archivo {ruta_origen}: {str(e)}")

def generar_y_enviar_excels_temporales(ruta_archivo):
    wb = load_workbook(ruta_archivo)
    gerentes_data = {}
    outlook = win32com.client.Dispatch("Outlook.Application")
    outlook.GetNamespace("MAPI").GetDefaultFolder(6).Display()
    for ws in wb.worksheets:
        data = pd.read_excel(ruta_archivo, sheet_name=ws.title)
        tipo_hoja = identificar_tipo_archivo(data)
        for gerente in data["CORREO GERENTE"].unique():
            if gerente is not None:
                if gerente not in gerentes_data:
                    gerentes_data[gerente] = {}
                data_filtrada = data[data['CORREO GERENTE'] == gerente]
                gerentes_data[gerente][tipo_hoja] = data_filtrada
    for gerente, datos in gerentes_data.items():
        nombre_gerente = gerente.split('@')[0].replace('.', ' ')
        ruta_archivo_gerente = os.path.join(RUTA_TEMPORAL, f"Transito pendiente {nombre_gerente}.xlsx")
        with pd.ExcelWriter(ruta_archivo_gerente, engine='openpyxl') as writer:
            for tipo_hoja, df in datos.items():
                df.to_excel(writer, sheet_name=tipo_hoja, index=False)
        wb_gerente = load_workbook(ruta_archivo_gerente)
        for ws in wb_gerente.worksheets:
            eliminar_columna(ws, "Columna1")
            eliminar_columna(ws, "Columna2")
            eliminar_columna(ws, "Columna3")
            eliminar_columna(ws, "CODIGO POS")
            eliminar_columna(ws, "CORREO GERENTE")
            max_col, max_row = ws.max_column, ws.max_row
            if max_row > 1:
                table_range = f"{ws.cell(row=1, column=1).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}"
                tabla = Table(displayName=f"Tabla_{ws.title.replace(' ', '_')}", ref=table_range)
                estilo = TableStyleInfo(name="TableStyleLight13", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                tabla.tableStyleInfo = estilo
                ws.add_table(tabla)
        wb_gerente.save(ruta_archivo_gerente)
        enviar_correo(outlook, gerente, ruta_archivo_gerente)
    archivos = [a for a in os.listdir(RUTA_CARPETA) if os.path.isfile(os.path.join(RUTA_CARPETA, a))]
    for archivo in archivos:
        ruta_completa_archivo = os.path.join(RUTA_CARPETA, archivo)
        print(ruta_completa_archivo)
        mover_archivo(ruta_completa_archivo, RUTA_HISTORICO)

def identificar_tipo_archivo(df):
    if "Pos Destino" in df.columns:
        return "Transferencias"
    elif "Pos" in df.columns:
        return "Compras"

def obtener_correos_gerentes():
    try:
        df_gerentes = pd.read_excel(RUTA_BD_GERENTES)
        if 'POS' not in df_gerentes.columns or 'CORREO CORPORATIVO' not in df_gerentes.columns:
            print("Error: No se encontraron las columnas 'POS' o 'CORREO CORPORATIVO'")
            return {}
        df_gerentes['POS'] = df_gerentes['POS'].fillna(0).astype(float).astype(int)
        df_gerentes['POS'] = df_gerentes['POS'].astype(str).str.strip()
        df_gerentes['POS'] = df_gerentes['POS'].apply(lambda x: x.zfill(4))
        df_gerentes = df_gerentes[df_gerentes['POS'].str.len() > 0]
        pos_to_correo = dict(zip(df_gerentes['POS'], df_gerentes['CORREO CORPORATIVO']))
        return pos_to_correo
    except Exception as e:
        print(f"Error al cargar base de datos de gerentes: {str(e)}")
        return {}

def extraer_codigo_pos(tienda):
    match = re.search(r'(\d{4})', tienda)
    return match.group(1) if match else None

def concatenar_bd(ruta_archivo):
    try:
        nombre_archivo = "Compras y Transferencias.xlsx"
        pos_to_correo = obtener_correos_gerentes()
        if not pos_to_correo:
            print("No se pudo obtener la información de correos de gerentes")
            return None
        ruta_concatenados = os.path.join(RUTA_CONCATENADO, nombre_archivo)
        df1 = pd.read_excel(ruta_archivo, sheet_name="COMPRAS", parse_dates=False)
        columnas_fecha_df1 = df1.select_dtypes(include=['datetime64']).columns
        for col in columnas_fecha_df1:
            df1[col] = df1[col].dt.strftime('%d/%m/%Y')
        df1['CODIGO POS'] = df1['Nombre Tienda'].apply(extraer_codigo_pos)
        df1['CODIGO POS'] = df1['CODIGO POS'].str.zfill(4)
        df1['CORREO GERENTE'] = df1['CODIGO POS'].map(pos_to_correo).fillna('mariana.giraldo@gco.com.co')
        df2 = pd.read_excel(ruta_archivo, sheet_name="TRANSFERENCIAS RESUMEN", parse_dates=False)
        columnas_fecha_df2 = df2.select_dtypes(include=["datetime64"]).columns
        for col in columnas_fecha_df2:
            df2[col] = df2[col].dt.strftime('%d/%m/%Y')
        df2['CODIGO POS'] = df2['Nombre Tienda'].apply(extraer_codigo_pos)
        df2['CODIGO POS'] = df2['CODIGO POS'].str.zfill(4)
        df2['CORREO GERENTE'] = df2['CODIGO POS'].map(pos_to_correo).fillna('mariana.giraldo@gco.com.co')
        with pd.ExcelWriter(ruta_concatenados, engine='openpyxl') as writer:
            df1.to_excel(writer, sheet_name="COMPRAS", index=False)
            df2.to_excel(writer, sheet_name="TRANSFERENCIAS RESUMEN", index=False)
        wb = load_workbook(ruta_concatenados)
        for ws in wb.worksheets:
            max_col, max_row = ws.max_column, ws.max_row
            table_range = f"{ws.cell(row=1, column=1).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}"
            tabla = Table(displayName=f"Tabla_{ws.title.replace(' ', '_')}", ref=table_range)
            estilo = TableStyleInfo(name="TableStyleLight13", showFirstColumn=False, showLastColumn=False, showColumnStripes=True, showRowStripes=True)
            tabla.tableStyleInfo = estilo
            ws.add_table(tabla)
        wb.save(ruta_concatenados)
        return ruta_concatenados
    except Exception as e:
        print(f"Error al concatenar archivos: {e}")
        return None

def obtener_ruta(ruta_carpeta):
    archivos = [os.path.join(ruta_carpeta, archivo) for archivo in os.listdir(ruta_carpeta) if os.path.isfile(os.path.join(ruta_carpeta, archivo))]
    ruta_archivo = archivos[0] if len(archivos) > 0 else None
    return ruta_archivo

def main():
    ruta_archivo = obtener_ruta(RUTA_CARPETA) #Solo viene un archivo
    print(ruta_archivo)
    if ruta_archivo:
        ruta_archivo_concatenado = concatenar_bd(ruta_archivo)
        generar_y_enviar_excels_temporales(ruta_archivo_concatenado)
    else:
        print("No se encontraron archivos en la carpeta...")

if __name__ == "__main__":
    main()